
import tkinter as tk
from tkinter import Tk, Toplevel, Label, Button, Entry, Frame, messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os

#creo los archivos de datos 
#BDD PARA USUARIOS
archivo_usuarios = "Datos_Usuarios.xlsx"
if os.path.exists(archivo_usuarios):
    wb = load_workbook(archivo_usuarios)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Usuario", "Contraseña", "Tipo"])
    wb.save(archivo_usuarios)

#BDD PARA MESAS/RESERVAS
archivo_Reservas = "Datos_Reservas.xlsx"
if os.path.exists(archivo_Reservas):
    reservas = load_workbook(archivo_Reservas)
    wreservas = reservas.active
else:
    reservas = Workbook()
    wreservas = reservas.active
    wreservas.append(["ID", "Usuario", "Estado", "Capacidad"])
    reservas.save(archivo_Reservas)


#Arranca el sistema
class SeleccionarTipo:
    def __init__(self, ventana):
        self.ventana= ventana
        self.ventana.geometry("400x700")
        self.ventana.title("Tipo de Usuario")

        fondo = "#88FFB4"
        #Dividimos en dos partes la pantalla

        self.parte_superior= Frame(self.ventana, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        self.parte_inferior= Frame(self.ventana, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        #Agregamos la imagen del restaurante 
        self.imagen = Image.open(r"Logo.png").resize((210, 220))
        self.render = ImageTk.PhotoImage(self.imagen)
        Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")

        # titulos
        Label(self.parte_superior, text="¿Quién eres?", font=("Calisto MT", 30, "bold"), bg=fondo).pack(pady=10)

        #Botones
        Button(self.parte_inferior, text="Cliente", width=35, height=2, font=("Arial", 14), command=lambda: self.elegir_cliente("Cliente")).grid(row=2, column=0, columnspan=2, padx=35, pady=5)

        Button(self.parte_inferior, text="Administrador", width=35, height=2, font=("Arial", 14), command=lambda: self.elegir_admin("Administrador")).grid(row=4, column=0, columnspan=4, padx=35, pady=5)

   
    def elegir_cliente(self,tipo):
        self.ventana.withdraw()
        Login(self.ventana, tipo)

    def elegir_admin(self,tipo):
        self.ventana.withdraw()
        Login_Ingresar(self.ventana, tipo)


#Clase para cliente donde ingresa o registra
class Login:
    def __init__(self, ventana, tipo): #Constructores
        self.tipo=tipo
        self.ventana=Toplevel(ventana)
        self.ventana.geometry("400x700")
        self.ventana.title("Seleccione el tipo de usuario")

        fondo = "#88FFB4"

        #Dividimos en dos partes la pantalla 
       
        self.parte_superior= Frame(self.ventana, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        self.parte_inferior= Frame(self.ventana, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        #Agregamos la imagen 
        self.imagen = Image.open(r"Logo.png").resize((210, 220))
        self.render = ImageTk.PhotoImage(self.imagen)
        Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")

        Label(self.parte_superior, text="Bienvenido", font=("Calisto MT", 30, "bold"), bg=fondo).pack(pady=10)

        Button(self.parte_inferior, text="Registrar", width=35, height=2, font=("Arial", 14), command=lambda: self.eleccion_registrar(ventana)).grid(row=2, column=0, columnspan=2, padx=35, pady=5)

        Button(self.parte_inferior, text="Ingresar", width=35, height=2, font=("Arial", 14), command=lambda: self.eleccion_ingresar(ventana)).grid(row=4, column=0, columnspan=4, padx=35, pady=5)

        Button(self.parte_inferior, text="Atras", width=35, height=2, font=("Arial", 14), command=lambda: self.regresar(ventana)).grid(row=6, column=0, columnspan=6, padx=35, pady=5)


         # --- BOTÓN DE INFORMACIÓN REDONDO ARRIBA A LA DERECHA ---
        self.canvas_info = tk.Canvas(self.parte_superior, width=30, height=30, highlightthickness=0, bg=fondo)
        self.canvas_info.place(relx=1.0, x=-15, y=15, anchor="ne")

        circulo = self.canvas_info.create_oval(2, 2, 28, 28, fill="#3498db", outline="#2980b9")
        texto = self.canvas_info.create_text(15, 15, text="i", fill="white", font=("Times New Roman", 14, "italic"))

        # hacer que tanto el círculo como la letra sean clickeables
        self.canvas_info.tag_bind(circulo, "<Button-1>", lambda e: self.mostrar_info())
        self.canvas_info.tag_bind(texto, "<Button-1>", lambda e: self.mostrar_info())

    def regresar(self,ventana):
        self.ventana.destroy()
        ventana.deiconify()
    
    def mostrar_info(self):
        ventana_info = Toplevel(self.ventana)
        ventana_info.title("Información")
        ventana_info.geometry("400x200")
        Label(ventana_info, text="Abrimos de jueves a domingo...\n\n Te acompañamos en tus momentos especiales,\n ya sea en el almuerzo o en la cena. ¡Te esperamos! \n\n\n Número de contacto: 3546 879736", wraplength=280, justify="center").pack(pady=20)
        Button(ventana_info, text="Cerrar", command=ventana_info.destroy).pack(pady=10)

        
    def eleccion_ingresar(self, ventana):
        self.ventana.destroy()
        Login_Ingresar(ventana,self.tipo)

    def eleccion_registrar(self, ventana):
        self.ventana.destroy()
        Registro(ventana,self.tipo)

#clase cuando se toma la opcion de registrar y si todo esta bien se va a la pestaña login
class Registro:
        def __init__(self, ventana, tipo): #Constructores
            self.tipo=tipo
            self.ventana= Toplevel(ventana)
            self.ventana.geometry("400x700")
            self.ventana.title("Registro")

            fondo = "#88FFB4"

            self.parte_superior = Frame(self.ventana, bg=fondo)
            self.parte_superior.pack(fill="both", expand=True)

            self.parte_inferior = Frame(self.ventana, bg=fondo)
            self.parte_inferior.pack(fill="both", expand=True)

            self.parte_inferior.columnconfigure(0, weight=1)
            self.parte_inferior.columnconfigure(1, weight=1)
          
            #titulo
            Label(self.parte_superior, text=f"Tipo: {self.tipo}", font=("Calisto MT", 30, "bold"), bg=fondo).pack(side="top", pady=20)

            #Imagen Registro
            self.imagen = Image.open(r"registrarse.png").resize((210, 220))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")

            #Entradas
            Label(self.parte_inferior, text="Usuario", font=("Arial", 18), bg=fondo).grid(row=0, column=0, padx=10, sticky="e")
            self.entry_usuario = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18))
            self.entry_usuario.grid(row=0, column=1, columnspan=3, padx=5, sticky="w")

            Label(self.parte_inferior, text="Contraseña:", font=("Arial", 18), bg=fondo).grid(row=1, column=0, pady=10, sticky="e")
            self.entry_contrasena = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18), show="*")
            self.entry_contrasena.grid(row=1, column=1, columnspan=3, padx=5, sticky="w")

            Button(self.parte_inferior, text="Guardar", width=16, font=("Arial", 12), command=lambda: self.guardar_datos(ventana)).grid(row=2, column=0, columnspan=2, padx=35, pady=5)

            Button(self.parte_inferior, text="Atras", width=16, font=("Arial", 12), command=lambda: self.regresar(ventana)).grid(row=8, column=0, columnspan=4, padx=35, pady=5)

        def regresar(self,ventana):
            self.ventana.destroy()
            Login(ventana,self.tipo)

        def guardar_datos(self, ventana):
            usuario = self.entry_usuario.get()
            contrasena = self.entry_contrasena.get()

            #se valida que los campos se hayan completado
            if not usuario or not contrasena:
                messagebox.showwarning("Cuidado!", "Se deben completar ambos campos")
                return

            #Agregamos usuario y contraseña al excel de usuarios
            ws.append([usuario, contrasena, self.tipo])
            wb.save(archivo_usuarios)
            messagebox.showinfo("Datos", "Se guardaron los datos")

            self.entry_usuario.delete(0, "end")
            self.entry_contrasena.delete(0, "end")

            self.ventana.destroy()
            Login_Ingresar(ventana, self.tipo)

class Login_Ingresar:
    def __init__(self, ventana, tipo): #Constructores
        self.tipo=tipo
        self.ventana= Toplevel(ventana)
        self.ventana.geometry("400x700")
        self.ventana.title("Ingresar")

        fondo = "#88FFB4"

        #dividimos
        self.parte_superior = Frame(self.ventana, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        self.parte_inferior = Frame(self.ventana, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        #Etiqueta
        Label(self.parte_superior, text="Inicio de sesión", font=("Calisto MT", 30, "bold"), bg=fondo).pack(side="top", pady=20)

        #Imagen
        self.imagen = Image.open(r"Imagen_Login.png").resize((210, 220))
        self.render = ImageTk.PhotoImage(self.imagen)
        Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        
        #Entradas
        Label(self.parte_inferior, text="Usuario", font=("Arial", 18), bg=fondo).grid(row=0, column=0, padx=10, sticky="e")
        self.entry_usuario = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18))
        self.entry_usuario.grid(row=0, column=1, columnspan=3, padx=5, sticky="w")

        Label(self.parte_inferior, text="Contraseña:", font=("Arial", 18), bg=fondo).grid(row=1, column=0, pady=10, sticky="e")
        self.entry_contrasena = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18), show="*")
        self.entry_contrasena.grid(row=1, column=1, columnspan=3, padx=5, sticky="w")

        Button(self.parte_inferior, text="Ingresar", width=16, font=("Arial", 12), command=self.verificacion_datos).grid(row=2, column=0, columnspan=2, padx=35, pady=5)
        
        Button(self.parte_inferior, text="Atras", width=16, font=("Arial", 12), command=lambda: self.regresar(ventana)).grid(row=8, column=0, columnspan=4, padx=35, pady=5)
    
    def regresar(self,ventana):
        self.ventana.destroy()
        ventana.deiconify()

    def verificacion_datos(self):
        usuario = self.entry_usuario.get()
        contrasena = self.entry_contrasena.get()

        if not usuario or not contrasena:
            messagebox.showwarning("Cuidado!", "Se deben completar ambos campos")
            return

        if not os.path.exists(archivo_usuarios):
            messagebox.showerror("Error", "No hay usuarios registrados")
            return
        
        if self.tipo == "Administrador" and usuario == "admin" and contrasena == "123":
            messagebox.showinfo("Éxito", "Bienvenido admin como Administrador")
            self.ventana.destroy()
            Restaurante(self.ventana.master, self.tipo, usuario)
            return

        wb_temp = load_workbook(archivo_usuarios)
        ws_temp = wb_temp.active

        encontrado = False
        for fila in ws_temp.iter_rows(values_only=True):
            if fila[0] == usuario and fila[1] == contrasena and fila[2] == self.tipo:
                encontrado = True
                break

        if encontrado:
            messagebox.showinfo("Éxito", f"Bienvenido {usuario} como {self.tipo}")
            self.ventana.destroy()
            Restaurante(self.ventana.master, self.tipo, usuario)
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrecto")
         
class Restaurante:
    def __init__(self, ventana, tipo, usuario):
        self.tipo = tipo
        self.usuario = usuario
        self.ventana = Toplevel(ventana)
        self.ventana.geometry("700x700")
        self.ventana.title("Restaurante")

        fondo = "#88FFB4"
        self.ventana.configure(bg=fondo)

        self.parte_superior = Frame(self.ventana, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        self.parte_inferior = Frame(self.ventana, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)

        # Botón de información
        self.canvas_info = tk.Canvas(self.parte_superior, width=30, height=30, highlightthickness=0, bg=fondo)
        self.canvas_info.place(relx=1.0, x=-15, y=15, anchor="ne")
        circulo = self.canvas_info.create_oval(2, 2, 28, 28, fill="#3498db", outline="#2980b9")
        texto = self.canvas_info.create_text(15, 15, text="i", fill="white", font=("Times New Roman", 14, "italic"))
        self.canvas_info.tag_bind(circulo, "<Button-1>", lambda e: self.mostrar_info())
        self.canvas_info.tag_bind(texto, "<Button-1>", lambda e: self.mostrar_info())

        # Imagen
        self.imagen = Image.open("Logo.png").resize((170, 180))
        self.render = ImageTk.PhotoImage(self.imagen)
        Label(self.parte_superior, image=self.render, bg=fondo).pack(pady=10)

        Button(self.parte_inferior, text="Cerrar Sesión", width=16, font=("Arial", 12),
               command=lambda: self.cerrar_sesion(ventana)).pack(pady=10)

        self.mesas = 6
        self.filas = []
        self.estado_mesas = {}
        self.botones_mesas = {}

        # Crear mesas iniciales
        for i in range(1, self.mesas + 1):
            if (i - 1) % 3 == 0:
                fila_frame = Frame(self.parte_superior, bg=fondo)
                fila_frame.pack()
                self.filas.append(fila_frame)

            btn = Button(
                self.filas[-1],
                text=f"Mesa {i}",
                width=12,
                height=3,
                bg="green",
                command=lambda n=i: Reservar_mesa(self.ventana, self.tipo, self, n)
            )
            btn.pack(side="left", padx=10, pady=10)
            self.estado_mesas[i] = {"estado": "Libre", "nombre": ""}
            self.botones_mesas[i] = btn

        # Cargar estado desde Excel
        archivo_Reservas = "Datos_Reservas.xlsx"
        if os.path.exists(archivo_Reservas):
            reservas = load_workbook(archivo_Reservas)
            hoja = reservas.active
            max_id = 0
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                # Validar que la fila tenga al menos 4 columnas y que mesa_id no sea None
                if fila is None or len(fila) < 4 or fila[0] is None:
                    continue
                mesa_id, nombre, estado, capacidad = fila[:4]
                try:
                    mesa_id = int(mesa_id)
                except (TypeError, ValueError):
                    continue
                max_id = max(max_id, mesa_id)
                if mesa_id not in self.botones_mesas:
                    self.mesas += 1
                    fila_index = (self.mesas - 1) // 3
                    if fila_index >= len(self.filas):
                        fila_frame = Frame(self.parte_superior, bg=fondo)
                        fila_frame.pack()
                        self.filas.append(fila_frame)
                    else:
                        fila_frame = self.filas[fila_index]
            
                    btn = Button(
                        fila_frame,
                        text=f"Mesa {mesa_id}\nCapacidad: {capacidad}",
                        width=12,
                        height=3,
                        bg="green",
                        command=lambda n=mesa_id: Reservar_mesa(self.ventana, self.tipo, self, n)
                    )
                    btn.pack(side="left", padx=10, pady=10)
                    self.botones_mesas[mesa_id] = btn
            
                self.estado_mesas[mesa_id] = {"estado": estado, "nombre": nombre, "capacidad": capacidad}
                color = "red" if estado == "Ocupada" else "green"
                texto = f"Mesa {mesa_id}\nCapacidad: {capacidad}" if estado == "Libre" else f"Mesa {mesa_id}\n({nombre})\nCapacidad: {capacidad}"
                self.botones_mesas[mesa_id].configure(bg=color, text=texto)

        if self.tipo == "Administrador":
            Button(self.parte_inferior, text="Agregar Mesa", width=16, font=("Arial", 12),
                   command=self.agregar_mesa).pack(pady=10)

    def cerrar_sesion(self, ventana):
        self.ventana.destroy()
        ventana.deiconify()

    def agregar_mesa(self):
            ventana_capacidad = Toplevel(self.ventana)
            ventana_capacidad.title("Capacidad de la Mesa")
            ventana_capacidad.geometry("300x150")
            
            Label(ventana_capacidad, text="Ingrese la capacidad:", font=("Arial", 14)).pack(pady=10)
            entry_capacidad = Entry(ventana_capacidad, font=("Arial", 14))
            entry_capacidad.pack(pady=5)

            def guardar_capacidad():
                capacidad = entry_capacidad.get()
                if not capacidad.isdigit() or int(capacidad) <= 0:
                    messagebox.showerror("Error", "Ingrese un número válido")
                    return

                self.mesas += 1
                fila_index = (self.mesas - 1) // 3

                if fila_index >= len(self.filas):
                    fila_frame = Frame(self.parte_superior, bg="#88FFB4")
                    fila_frame.pack()
                    self.filas.append(fila_frame)
                else:
                    fila_frame = self.filas[fila_index]

                btn = Button(
                    fila_frame,
                    text=f"Mesa {self.mesas}\nCapacidad: {capacidad}",
                    width=12,
                    height=3,
                    bg="green",
                    command=lambda n=self.mesas: Reservar_mesa(self.ventana, self.tipo, self, n)
                )
                btn.pack(side="left", padx=10, pady=10)
                self.estado_mesas[self.mesas] = {"estado": "Libre", "nombre": "", "capacidad": capacidad}
                self.botones_mesas[self.mesas] = btn

                archivo_Reservas = "Datos_Reservas.xlsx"
                reservas = load_workbook(archivo_Reservas)
                hoja = reservas.active
                hoja.append([self.mesas, "", "Libre", capacidad])
                reservas.save(archivo_Reservas)

                ventana_capacidad.destroy()

            Button(ventana_capacidad, text="Guardar", command=guardar_capacidad).pack(pady=10)

    def mostrar_info(self):
        ventana_info = Toplevel(self.ventana)
        ventana_info.title("Nuestra Historia")
        ventana_info.geometry("400x250")
        Label(
            ventana_info,
            text="Hace más de 20 años abrimos nuestras puertas\n"
                 "con la idea de compartir el sabor de la cocina casera.\n\n"
                 "Con el tiempo nos convertimos en un punto de encuentro\n"
                 "para familias y amigos que buscan disfrutar\n"
                 "de buena comida y momentos inolvidables.",
            wraplength=320,
            justify="center"
        ).pack(pady=20)
        Button(ventana_info, text="Cerrar", command=ventana_info.destroy).pack(pady=10)

class Reservar_mesa:
    def __init__(self, ventana, tipo, restaurante, mesa_id):
        self.tipo = tipo
        self.restaurante = restaurante
        self.mesa_id = mesa_id

        self.ventana = Toplevel(ventana)
        self.ventana.geometry("400x700")
        self.ventana.title("Reservar Mesa")

        fondo = "#88FFB4"

        self.parte_superior = Frame(self.ventana, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        self.parte_inferior = Frame(self.ventana, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)
        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        # Recuperar capacidad
        capacidad = self.restaurante.estado_mesas[self.mesa_id].get("capacidad", "4")
        titulo = (
            f"Admin - Mesa {mesa_id} (Capacidad: {capacidad})"
            if self.tipo == "Administrador"
            else f"Reservar Mesa {mesa_id} (Capacidad: {capacidad})"
        )
        Label(self.parte_superior, text=titulo, font=("Calisto MT", 20, "bold"), bg=fondo).pack(pady=20)

        # Imagen
        try:
            self.imagen = Image.open("reserva.png").resize((210, 220))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        except FileNotFoundError:
            Label(self.parte_superior, text="[Imagen no encontrada]", bg=fondo).pack()

        mesa_info = self.restaurante.estado_mesas[self.mesa_id]

        # Restricción si está ocupada
        if mesa_info["estado"] == "Ocupada" and self.tipo != "Administrador" and mesa_info["nombre"] != self.restaurante.usuario:
            Label(
                self.parte_inferior,
                text=f"Mesa ocupada por {mesa_info['nombre']}",
                font=("Arial", 16), bg=fondo
            ).grid(row=0, column=0, columnspan=2, pady=10)
            Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),
                   command=self.regresar).grid(row=2, column=0, columnspan=2, pady=5)
            return

        # Botones según el tipo de usuario
        if self.tipo == "Administrador":
            Button(self.parte_inferior, text="Desocupar", width=16, font=("Arial", 12),
                   command=self.desocupar).grid(row=2, column=0, columnspan=2, pady=5)
            Button(self.parte_inferior, text="Eliminar", width=16, font=("Arial", 12),
                   command=self.eliminar).grid(row=4, column=0, columnspan=2, pady=5)
            Button(self.parte_inferior, text="Modificar Capacidad", width=16, font=("Arial", 12),
                   command=self.modificar_capacidad).grid(row=5, column=0, columnspan=2, pady=5)
            Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),
                   command=self.regresar).grid(row=6, column=0, columnspan=2, pady=5)

        else:
            Label(self.parte_inferior, text="Nombre:", font=("Arial", 16), bg=fondo).grid(row=0, column=0, pady=10, sticky="e")
            self.entry_nombre = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 16))
            self.entry_nombre.grid(row=0, column=1, padx=5, sticky="w")

            Button(self.parte_inferior, text="Guardar", width=16, font=("Arial", 12),
                   command=self.guardar).grid(row=2, column=0, columnspan=2, pady=5)
            Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),
                   command=self.regresar).grid(row=4, column=0, columnspan=2, pady=5)

        # Mostrar botón cancelar si corresponde
        self.mostrar_cancelar()

    # ---------------- MÉTODOS ---------------- #

    def regresar(self):
        self.ventana.destroy()

    def guardar(self):
        """Guardar la reserva hecha por un usuario"""
        nombre = self.restaurante.usuario  # Usar el usuario logueado como nombre de reserva
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["estado"] == "Libre":
            capacidad = mesa_info.get("capacidad", "4")
            self.restaurante.estado_mesas[self.mesa_id] = {"estado": "Ocupada", "nombre": nombre, "capacidad": capacidad}
            self.restaurante.botones_mesas[self.mesa_id].configure(
                bg="red", text=f"Mesa {self.mesa_id}\n({nombre})\nCapacidad: {capacidad}"
            )
            self.actualizar_excel("Ocupada", nombre, capacidad)
            self.ventana.destroy()
        else:
            messagebox.showerror("Error", "Esta mesa ya está ocupada")

    def modificar_capacidad(self):
        ventana_cap = Toplevel(self.ventana)
        ventana_cap.title("Modificar Capacidad")
        ventana_cap.geometry("300x150")

        Label(ventana_cap, text="Nueva capacidad:", font=("Arial", 14)).pack(pady=10)
        entry_cap = Entry(ventana_cap, font=("Arial", 14))
        entry_cap.pack(pady=5)

        def guardar():
            nueva_cap = entry_cap.get()
            if not nueva_cap.isdigit() or int(nueva_cap) <= 0:
                messagebox.showerror("Error", "Ingrese un número válido")
                return

            mesa_info = self.restaurante.estado_mesas[self.mesa_id]
            mesa_info["capacidad"] = nueva_cap

            texto = f"Mesa {self.mesa_id}\nCapacidad: {nueva_cap}" if mesa_info["estado"] == "Libre" else f"Mesa {self.mesa_id}\n({mesa_info['nombre']})\nCapacidad: {nueva_cap}"
            self.restaurante.botones_mesas[self.mesa_id].configure(text=texto)

            self.actualizar_excel(mesa_info["estado"], mesa_info["nombre"], nueva_cap)
            ventana_cap.destroy()

        Button(ventana_cap, text="Guardar", command=guardar).pack(pady=10)

    def eliminar(self):
        if messagebox.askyesno("Eliminar Mesa", "¿Está seguro que desea eliminar esta mesa?"):
            self.eliminar_de_excel()
            btn = self.restaurante.botones_mesas.get(self.mesa_id)
            if btn:
                btn.destroy()
                del self.restaurante.botones_mesas[self.mesa_id]
            self.restaurante.estado_mesas.pop(self.mesa_id, None)
            self.ventana.destroy()

    def desocupar(self):
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["nombre"] == self.restaurante.usuario or self.tipo == "Administrador":
            capacidad = mesa_info.get("capacidad", "4")
            self.restaurante.estado_mesas[self.mesa_id] = {"estado": "Libre", "nombre": "", "capacidad": capacidad}
            self.restaurante.botones_mesas[self.mesa_id].configure(
                bg="green", text=f"Mesa {self.mesa_id}\nCapacidad: {capacidad}"
            )
            self.actualizar_excel("Libre", "", capacidad)
            self.ventana.destroy()
        else:
            messagebox.showerror("Error", "Solo el administrador o el usuario que reservó puede desocupar la mesa.")

    def actualizar_excel(self, estado, nombre="", capacidad="4"):
        archivo = "Datos_Reservas.xlsx"
        reservas = load_workbook(archivo)
        hoja = reservas.active

        for fila in hoja.iter_rows(min_row=2):
            if fila[0].value == self.mesa_id:
                fila[1].value = nombre
                fila[2].value = estado
                fila[3].value = capacidad
                reservas.save(archivo)
                return

        hoja.append([self.mesa_id, nombre, estado, capacidad])
        reservas.save(archivo)

    def cancelar_reserva(self):
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["estado"] == "Ocupada" and mesa_info["nombre"] == self.restaurante.usuario:
            capacidad = mesa_info.get("capacidad", "4")
            self.restaurante.estado_mesas[self.mesa_id] = {"estado": "Libre", "nombre": "", "capacidad": capacidad}
            self.restaurante.botones_mesas[self.mesa_id].configure(
                bg="green", text=f"Mesa {self.mesa_id}\nCapacidad: {capacidad}"
            )
            self.actualizar_excel("Libre", "", capacidad)
            self.ventana.destroy()
        else:
            messagebox.showerror("Error", "Solo el usuario que reservó puede cancelar la reserva.")

    def mostrar_cancelar(self):
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["estado"] == "Ocupada" and mesa_info["nombre"] == self.restaurante.usuario and self.tipo != "Administrador":
            Button(self.parte_inferior, text="Cancelar Reserva", width=16, font=("Arial", 12),
                   command=self.cancelar_reserva).grid(row=6, column=0, columnspan=2, pady=5)

    def eliminar_de_excel(self):
        archivo = "Datos_Reservas.xlsx"
        reservas = load_workbook(archivo)
        hoja = reservas.active

        for fila in hoja.iter_rows(min_row=2):
            if fila[0].value == self.mesa_id:
                hoja.delete_rows(fila[0].row)
                break

        reservas.save(archivo)


ventana= Tk()
aplicacion=SeleccionarTipo(ventana)
ventana.mainloop()

