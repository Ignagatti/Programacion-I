import tkinter as tk
from tkinter import Toplevel, Frame, Button, Label, Entry, messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os
from reservas import Reservar_mesa

archivo_Reservas = "Datos_Reservas.xlsx"
if os.path.exists(archivo_Reservas):
    reservas = load_workbook(archivo_Reservas)
    wreservas = reservas.active
else:
    # Si no existe el archivo, lo creo y agrego las 9 mesas por defecto
    reservas = Workbook()
    wreservas = reservas.active
    wreservas.append(["ID", "Usuario", "Estado", "Capacidad"])
    for i in range(1, 13):  # mesas 1 a 12
        wreservas.append([i, "", "Libre", 4])  # capacidad por defecto 4
    reservas.save(archivo_Reservas)


class Restaurante:
    def __init__(self, ventana, tipo, usuario):
        # Constructores
        self.tipo = tipo
        self.usuario = usuario
        self.ventana = Toplevel(ventana)
        self.ventana.geometry("900x700+300+100")
        self.ventana.title("Restaurante")

        fondo = "#427c64"
        fondo_derecha = "#e5e8df"
        color_botones = "#A7CBBF"

        # CONTENEDOR PRINCIPAL DIVIDIDO EN 2 COLUMNAS
        self.contenedor = Frame(self.ventana, bg=fondo)
        self.contenedor.pack(fill="both", expand=True)

        self.contenedor.columnconfigure(0, weight=3)  # izquierda = mesas
        self.contenedor.columnconfigure(1, weight=2)  # derecha = formulario
        self.contenedor.rowconfigure(0, weight=1)

        # FRAME IZQUIERDO (MESAS)
        self.izquierda = Frame(self.contenedor, bg=fondo, relief="raised", bd=6)
        self.izquierda.grid(row=0, column=0, sticky="nsew")

        # FRAME DERECHO (FORMULARIO)
        self.derecha = Frame(self.contenedor, bg=fondo_derecha, relief="raised", bd=4)
        self.derecha.grid(row=0, column=1, sticky="nsew")

        # Parte superior: título e imagen de login
        self.parte_superior = Frame(self.derecha, bg=fondo_derecha)
        self.parte_superior.pack(fill="both", expand=True)

        Label(
            self.parte_superior,
            text="Que desea hacer?",
            font=("Calisto MT", 20, "bold"),
            bg=fondo_derecha,
        ).pack(side="top", pady=20)

        # Imagen
        try:
            self.imagen = Image.open(r"Logo.png").resize((180, 190))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.parte_superior, image=self.render, bg=fondo_derecha).pack(
                expand=True, fill="both", side="top"
            )
        except Exception:
            Label(
                self.parte_superior,
                text="[Imagen login no encontrada]",
                bg=fondo_derecha,
                font=("Arial", 14, "italic"),
                fg="gray",
            ).pack(expand=True, fill="both", side="top")

        # Parte inferior: entradas y botones
        self.parte_inferior = Frame(self.derecha, bg=fondo_derecha)
        self.parte_inferior.pack(fill="both", expand=True)

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        # Botón de información
        self.canvas_info = tk.Canvas(
            self.izquierda, width=30, height=30, highlightthickness=0, bg=fondo
        )
        self.canvas_info.place(relx=1.0, x=-15, y=15, anchor="ne")
        circulo = self.canvas_info.create_oval(
            2, 2, 28, 28, fill="#3498db", outline="#2980b9"
        )
        texto = self.canvas_info.create_text(
            15, 15, text="i", fill="white", font=("Times New Roman", 14, "italic")
        )
        self.canvas_info.tag_bind(circulo, "<Button-1>", lambda e: self.mostrar_info())
        self.canvas_info.tag_bind(texto, "<Button-1>", lambda e: self.mostrar_info())

        # Diccionarios de control
        self.estado_mesas = {}
        self.botones_mesas = {}

        # Cargar estado desde Excel (siempre)
        reservas = load_workbook(archivo_Reservas)
        hoja = reservas.active
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila is None or len(fila) < 4 or fila[0] is None:
                continue
            mesa_id, nombre, estado, capacidad = fila[:4]
            try:
                mesa_id = int(mesa_id)
            except (TypeError, ValueError):
                continue

            texto_btn = (
                f"Mesa {mesa_id}\nCapacidad: {capacidad}"
                if estado == "Libre"
                else f"Mesa {mesa_id}\n({nombre})\nCapacidad: {capacidad}"
            )

            fila_index = (mesa_id - 1) // 3
            columna_index = (mesa_id - 1) % 3

            btn = Button(
                self.izquierda,
                text=texto_btn,
                width=12,
                height=3,
                bg="red" if estado == "Ocupada" else "green",
                relief="raised", bd=4, command=lambda n=mesa_id: Reservar_mesa(
                    self.ventana, self.tipo, self, n
                ),
            )
            btn.grid(row=fila_index, column=columna_index, padx=40, pady=20)

            self.estado_mesas[mesa_id] = {
                "estado": estado,
                "nombre": nombre,
                "capacidad": capacidad,
            }
            self.botones_mesas[mesa_id] = btn

        # Botones solo para administradores
        if self.tipo == "Administrador":
            Button(
                self.parte_inferior,
                text="Agregar Mesa",
                width=16,
                font=("Arial", 12),
                bg=color_botones,
                relief="raised",command=self.agregar_mesa,
            ).pack(pady=10)

        Button(
            self.parte_inferior,
            text="Cerrar Sesión",
            width=16,
            font=("Arial", 12),
            bg=color_botones,
            relief="raised",command=lambda: self.cerrar_sesion(ventana),
        ).pack(pady=10)

    def cerrar_sesion(self, ventana):
        self.ventana.destroy()
        ventana.deiconify()

    def agregar_mesa(self):
        ventana_capacidad = Toplevel(self.ventana)
        ventana_capacidad.title("Capacidad de la Mesa")
        ventana_capacidad.geometry("300x150")

        Label(ventana_capacidad, text="Ingrese la capacidad:", font=("Arial", 14)).pack(
            pady=10
        )
        entry_capacidad = Entry(ventana_capacidad, font=("Arial", 14))
        entry_capacidad.pack(pady=5)

        def guardar_capacidad():
            capacidad = entry_capacidad.get()
            if not capacidad.isdigit() or int(capacidad) <= 0:
                messagebox.showerror("Error", "Ingrese un número válido")
                return

            reservas = load_workbook(archivo_Reservas)
            hoja = reservas.active

            # Buscar huecos en los IDs
            ids_existentes = [
                fila[0].value for fila in hoja.iter_rows(min_row=2) if fila[0].value
            ]
            ids_existentes = sorted(ids_existentes)

            nuevo_id = None
            for i in range(1, max(ids_existentes) + 2):  # busca el primer hueco
                if i not in ids_existentes:
                    nuevo_id = i
                    break

            if nuevo_id is None:
                nuevo_id = max(ids_existentes) + 1

            fila_index = (nuevo_id - 1) // 3
            columna_index = (nuevo_id - 1) % 3

            btn = Button(
                self.izquierda,
                text=f"Mesa {nuevo_id}\nCapacidad: {capacidad}",
                width=12,
                height=3,
                bg="green",
                relief="raised", bd=4,command=lambda n=nuevo_id: Reservar_mesa(
                    self.ventana, self.tipo, self, n
                ),
            )
            btn.grid(row=fila_index, column=columna_index, padx=40, pady=20)
            self.estado_mesas[nuevo_id] = {
                "estado": "Libre",
                "nombre": "",
                "capacidad": capacidad,
            }
            self.botones_mesas[nuevo_id] = btn

            # Guardar en Excel
            hoja.append([nuevo_id, "", "Libre", capacidad])
            reservas.save(archivo_Reservas)

            ventana_capacidad.destroy()

        Button(ventana_capacidad, text="Guardar", command=guardar_capacidad).pack(
            pady=10
        )

   
    def mostrar_info(self):
        ventana_info = Toplevel(self.ventana)
        ventana_info.title("Nuestra Historia")
        ventana_info.geometry("400x250+300+100")
        Label(
            ventana_info,
            text="Hace más de 20 años abrimos nuestras puertas\n"
            "con la idea de compartir el sabor de la cocina casera.\n\n"
            "Con el tiempo nos convertimos en un punto de encuentro\n"
            "para familias y amigos que buscan disfrutar\n"
            "de buena comida y momentos inolvidables.",
            wraplength=320,
            justify="center",
        ).pack(pady=20)
        Button(ventana_info, text="Cerrar", command=ventana_info.destroy).pack(pady=10)