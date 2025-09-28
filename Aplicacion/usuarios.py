# usuarios.py
import tkinter as tk
from tkinter import Toplevel, Label, Button, Entry, Frame, messagebox
from PIL import Image, ImageTk
from restaurante import Restaurante
from openpyxl import Workbook, load_workbook
import os

archivo_usuarios = "Datos_Usuarios.xlsx"
if os.path.exists(archivo_usuarios):
    wb = load_workbook(archivo_usuarios)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Usuario", "Contraseña", "Tipo"])
    wb.save(archivo_usuarios)

class SeleccionarTipo:
    def __init__(self, ventana):
        self.ventana = ventana
        self.ventana.geometry("800x600")
        self.ventana.title("Tipo de Usuario")

        fondo = "#588E6B"

        # Frame principal
        self.contenedor = Frame(self.ventana, bg=fondo)
        self.contenedor.pack(fill="both", expand=True)

        # Configuración de columnas y filas
        self.contenedor.columnconfigure(0, weight=1)  # izquierda
        self.contenedor.columnconfigure(1, weight=1)  # derecha
        self.contenedor.rowconfigure(0, weight=1)     # superior
        self.contenedor.rowconfigure(1, weight=1)     # inferior

        #Frame izquierdo
        self.izquierda = Frame(self.contenedor, bg=fondo)
        self.izquierda.grid(row=0, column=0, rowspan=2, sticky="nsew")

        #Frame superior derecho
        self.der_superior = Frame(self.contenedor, bg=fondo)
        self.der_superior.grid(row=0, column=1, sticky="nsew")

        #Frame inferior derecho (botones)
        self.der_inferior = Frame(self.contenedor, bg=fondo)
        self.der_inferior.grid(row=1, column=1, sticky="nsew")

        # Imagen ocupa todo el costado izquierdo
        try:
            self.original = Image.open("restaurante.png")
            self.img = ImageTk.PhotoImage(self.original.resize((400, 600)))
            self.label_imagen = Label(self.izquierda, image=self.img, bg=fondo)
        except Exception:
            self.label_imagen = Label(self.izquierda, text="[Imagen restaurante no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray")
        self.label_imagen.pack(fill="both", expand=True)

        # Redimensionar imagen automáticamente al cambiar el tamaño de la ventana
        self.izquierda.bind("<Configure>", self.redimensionar_imagen)

        #imagen logo restautrante
        try:
            self.imagen = Image.open(r"Logo.png").resize((210, 220))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.der_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        except Exception:
            Label(self.der_superior, text="[Logo no encontrado]", bg=fondo, font=("Arial", 14, "italic"), fg="gray").pack(expand=True, fill="both", side="top")
        
        # titulo
        Label(self.der_superior, text="¿Quién eres?", font=("Calisto MT", 30, "bold"), bg=fondo).pack(pady=10)
       
        color_botones= "#A7CBBF"
        # botones
        self.der_inferior.columnconfigure(0, weight=1)

        Button(self.der_inferior, text="Cliente", width=25, height=2, font=("Arial", 14), bg=color_botones, command=lambda: self.elegir_cliente("Cliente"),
                relief="raised", bd=4).grid(row=0, column=0, pady=10, padx=20)

        Button(self.der_inferior, text="Administrador", width=25, height=2, font=("Arial", 14), bg=color_botones, command=lambda: self.elegir_admin("Administrador"),
                relief="raised", bd=4).grid(row=1, column=0, pady=10, padx=20)

    def redimensionar_imagen(self, event):
        # Ajusta la imagen al tamaño del frame izquierdo
        nueva_img = self.original.resize((event.width, event.height))
        self.img = ImageTk.PhotoImage(nueva_img)
        self.label_imagen.config(image=self.img)

    def elegir_cliente(self, tipo):
        self.ventana.withdraw()
        Login(self.ventana, tipo)

    def elegir_admin(self, tipo):
        self.ventana.withdraw()
        Login_Ingresar(self.ventana, tipo)

class Login:
    def __init__(self, ventana, tipo): #Constructores
        self.tipo=tipo
        self.ventana=Toplevel(ventana)
        self.ventana.geometry("800x600")  # ahora más ancho para que entre la imagen
        self.ventana.title("Seleccione el tipo de usuario")

        fondo = "#588E6B"
        color_botones= "#A7CBBF"

        #DIVIDIMOS LA PANTALLA 
        self.contenedor = Frame(self.ventana, bg=fondo)
        self.contenedor.pack(fill="both", expand=True)

        self.contenedor.columnconfigure(0, weight=1)  # lado izquierdo
        self.contenedor.columnconfigure(1, weight=1)  # lado derecho
        self.contenedor.rowconfigure(0, weight=1)     # arriba
        self.contenedor.rowconfigure(1, weight=1)     # abajo

        #FRAME IZQUIERDO (imagen)
        self.izquierda = Frame(self.contenedor, bg=fondo)
        self.izquierda.grid(row=0, column=0, rowspan=2, sticky="nsew")

        # Imagen ocupa todo el costado izquierdo
        try:
            self.original = Image.open("restaurante.png")
            self.img = ImageTk.PhotoImage(self.original.resize((400, 600)))
            self.label_imagen = Label(self.izquierda, image=self.img, bg=fondo)
        except Exception:
            self.label_imagen = Label(self.izquierda, text="[Imagen restaurante no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray")
        self.label_imagen.pack(fill="both", expand=True)

        # Redimensionar imagen automáticamente al cambiar el tamaño de la ventana
        self.izquierda.bind("<Configure>", self.redimensionar_imagen)

        #FRAME DERECHO SUPERIOR 
        self.parte_superior= Frame(self.contenedor, bg=fondo)
        self.parte_superior.grid(row=0, column=1, sticky="nsew")

        #Agregamos la imagen 
        try:
            self.imagen = Image.open(r"Logo.png").resize((210, 220))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        except Exception:
            Label(self.parte_superior, text="[Logo no encontrado]", bg=fondo, font=("Arial", 14, "italic"), fg="gray").pack(expand=True, fill="both", side="top")

        Label(self.parte_superior, text="Bienvenido", font=("Calisto MT", 30, "bold"), bg=fondo).pack(pady=10)

        #BOTÓN DE INFORMACIÓN REDONDO ARRIBA A LA DERECHA 
        self.canvas_info = tk.Canvas(self.parte_superior, width=30, height=30, highlightthickness=0, bg=fondo)
        self.canvas_info.place(relx=1.0, x=-15, y=15, anchor="ne")

        circulo = self.canvas_info.create_oval(2, 2, 28, 28, fill="#3498db", outline="#2980b9")
        texto = self.canvas_info.create_text(15, 15, text="i", fill="white", font=("Times New Roman", 14, "italic"))

        # hacer que tanto el círculo como la letra sean clickeables
        self.canvas_info.tag_bind(circulo, "<Button-1>", lambda e: self.mostrar_info())
        self.canvas_info.tag_bind(texto, "<Button-1>", lambda e: self.mostrar_info())

        #FRAME DERECHO INFERIOR (botones)
        self.parte_inferior= Frame(self.contenedor, bg=fondo)
        self.parte_inferior.grid(row=1, column=1, sticky="nsew")

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        Button(self.parte_inferior, text="Registrar", width=35, height=2, font=("Arial", 14), bg= color_botones, command=lambda: self.eleccion_registrar(ventana), relief="raised", bd=4).grid(row=2, column=0, columnspan=2, padx=35, pady=5)

        Button(self.parte_inferior, text="Ingresar", width=35, height=2, font=("Arial", 14), bg= color_botones, command=lambda: self.eleccion_ingresar(ventana), relief="raised", bd=4).grid(row=4, column=0, columnspan=4, padx=35, pady=5)

        Button(self.parte_inferior, text="Atras", width=35, height=2, font=("Arial", 14), bg=color_botones, command=lambda: self.regresar(ventana), relief="raised", bd=4).grid(row=6, column=0, columnspan=6, padx=35, pady=5)

    def redimensionar_imagen(self, event):
        # Ajusta la imagen al tamaño del frame izquierdo
        nueva_img = self.original.resize((event.width, event.height))
        self.img = ImageTk.PhotoImage(nueva_img)
        self.label_imagen.config(image=self.img)

    def regresar(self,ventana):
        self.ventana.destroy()
        ventana.deiconify()
    
    def mostrar_info(self):
        ventana_info = Toplevel(self.ventana)
        ventana_info.title("Información")
        ventana_info.geometry("400x200")
        Label(ventana_info, text="Abrimos de jueves a domingo...\n\n Te acompañamos en tus momentos especiales,\n ya sea en el almuerzo o en la cena. ¡Te esperamos! \n\n\n Número de contacto: 3546 879736", wraplength=280, justify="center").pack(pady=20)
        Button(ventana_info, text="Cerrar", command=ventana_info.destroy).pack(pady=10)
    
    def eleccion_ingresar(self, ventana):
        self.ventana.destroy()
        Login_Ingresar(ventana,self.tipo)

    def eleccion_registrar(self, ventana):
        self.ventana.destroy()
        Registro(ventana,self.tipo)


#clase cuando se toma la opcion de registrar y si todo esta bien se va a la pestaña login
class Registro:
    def __init__(self, ventana, tipo): #Constructores
        self.tipo = tipo
        self.ventana = Toplevel(ventana)
        self.ventana.geometry("800x600")  # más ancho para dividir en dos
        self.ventana.title("Registro")

        fondo = "#588E6B"
        color_botones= "#A7CBBF"

        #CONTENEDOR PRINCIPAL DIVIDIDO EN 2 COLUMNAS
        self.contenedor = Frame(self.ventana, bg=fondo)
        self.contenedor.pack(fill="both", expand=True)

        self.contenedor.columnconfigure(0, weight=3)  # lado izquierdo (imagen más grande)
        self.contenedor.columnconfigure(1, weight=2)  # lado derecho (formulario)
        self.contenedor.rowconfigure(0, weight=1)

        #FRAME IZQUIERDO (IMAGEN GRANDE)
        self.izquierda = Frame(self.contenedor, bg=fondo)
        self.izquierda.grid(row=0, column=0, sticky="nsew")

        try:
            self.original = Image.open("restaurante.png")
            self.img = ImageTk.PhotoImage(self.original.resize((400, 600)))
            self.label_imagen = Label(self.izquierda, image=self.img, bg=fondo)
        except Exception:
            self.label_imagen = Label(self.izquierda, text="[Imagen restaurante no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray")
        self.label_imagen.pack(fill="both", expand=True)

        # Redimensionar automáticamente
        self.izquierda.bind("<Configure>", self.redimensionar_imagen)

        #FRAME DERECHO (titulo, formulario)
        self.derecha = Frame(self.contenedor, bg=fondo)
        self.derecha.grid(row=0, column=1, sticky="nsew")

        # Parte superior
        self.parte_superior = Frame(self.derecha, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        Label(self.parte_superior, text=f"Tipo: {self.tipo}", font=("Calisto MT", 30, "bold"), bg=fondo).pack(side="top", pady=20)

        # Imagen Registro
        try:
            self.imagen = Image.open(r"registrarse.png").resize((210, 220))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        except Exception:
            Label(self.parte_superior, text="[Imagen registro no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray").pack(expand=True, fill="both", side="top")

        # Parte inferior (formulario)
        self.parte_inferior = Frame(self.derecha, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        # Entradas
        Label(self.parte_inferior, text="Usuario", font=("Arial", 18), bg=fondo).grid(row=0, column=0, padx=10, sticky="e")
        self.entry_usuario = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18))
        self.entry_usuario.grid(row=0, column=1, columnspan=3, padx=5, sticky="w")

        Label(self.parte_inferior, text="Contraseña:", font=("Arial", 18), bg=fondo).grid(row=1, column=0, pady=10, sticky="e")
        self.entry_contrasena = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18), show="*")
        self.entry_contrasena.grid(row=1, column=1, columnspan=3, padx=5, sticky="w")

        # Botones
        Button(self.parte_inferior, text="Guardar", width=16, font=("Arial", 12), bg=color_botones, command=lambda: self.guardar_datos(ventana), relief="raised", bd=4).grid(row=2, column=0, columnspan=2, padx=35, pady=5)

        Button(self.parte_inferior, text="Atras", width=16, font=("Arial", 12),bg=color_botones, command=lambda: self.regresar(ventana), relief="raised", bd=4).grid(row=8, column=0, columnspan=4, padx=35, pady=5)

    #Método para redimensionar la imagen izquierda
    def redimensionar_imagen(self, event):
        nueva_img = self.original.resize((event.width, event.height))
        self.img = ImageTk.PhotoImage(nueva_img)
        self.label_imagen.config(image=self.img)

    def regresar(self, ventana):
        self.ventana.destroy()
        Login(ventana, self.tipo)

    def guardar_datos(self, ventana):
        usuario = self.entry_usuario.get().strip()
        contrasena = self.entry_contrasena.get().strip()

        # se valida que los campos se hayan completado
        if not usuario or not contrasena:
            messagebox.showwarning("Cuidado!", "Se deben completar ambos campos")
            return

        # Si no existe el archivo, lo creo con encabezado
        if not os.path.exists(archivo_usuarios):
            wb = Workbook()
            ws = wb.active
            ws.append(["Usuario", "Contraseña", "Tipo"])
            wb.save(archivo_usuarios)

        wb = load_workbook(archivo_usuarios)
        ws = wb.active

        # Verificar si ya existe ese usuario
        for nombre in ws.iter_rows(min_row=2, values_only=True):
            if nombre[0] == usuario:
                messagebox.showerror("Error", "Ese nombre de usuario ya está registrado, pruebe con otro")
                wb.close()
                return

        # Si no existe, lo guardamos
        ws.append([usuario, contrasena, self.tipo])
        wb.save(archivo_usuarios)
        wb.close()

        messagebox.showinfo("Datos", "Se guardaron los datos")

        self.entry_usuario.delete(0, "end")
        self.entry_contrasena.delete(0, "end")

        self.ventana.destroy()
        Login_Ingresar(ventana, self.tipo)

class Login_Ingresar:
    def __init__(self, ventana, tipo):  # Constructores
        self.tipo = tipo
        self.ventana = Toplevel(ventana)
        self.ventana.geometry("800x600")  # más ancho para dividir en dos columnas
        self.ventana.title("Ingresar")

        fondo = "#588E6B"
        color_botones= "#A7CBBF"

        #CONTENEDOR PRINCIPAL DIVIDIDO EN 2 COLUMNAS
        self.contenedor = Frame(self.ventana, bg=fondo)
        self.contenedor.pack(fill="both", expand=True)

        self.contenedor.columnconfigure(0, weight=3)  # izquierda = imagen grande
        self.contenedor.columnconfigure(1, weight=2)  # derecha = formulario
        self.contenedor.rowconfigure(0, weight=1)

        #FRAME IZQUIERDO (IMAGEN GRANDE)
        self.izquierda = Frame(self.contenedor, bg=fondo)
        self.izquierda.grid(row=0, column=0, sticky="nsew")

        try:
            self.original = Image.open("restaurante.png")
            self.img = ImageTk.PhotoImage(self.original.resize((400, 600)))
            self.label_imagen = Label(self.izquierda, image=self.img, bg=fondo)
        except Exception:
            self.label_imagen = Label(self.izquierda, text="[Imagen restaurante no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray")
        self.label_imagen.pack(fill="both", expand=True)

        # Redimensionar automáticamente
        self.izquierda.bind("<Configure>", self.redimensionar_imagen)

        #FRAME DERECHO (FORMULARIO)
        self.derecha = Frame(self.contenedor, bg=fondo)
        self.derecha.grid(row=0, column=1, sticky="nsew")

        # Parte superior: título e imagen de login
        self.parte_superior = Frame(self.derecha, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        Label(self.parte_superior, text="Inicio de sesión", font=("Calisto MT", 30, "bold"), bg=fondo).pack(side="top", pady=20)

        try:
            self.imagen = Image.open(r"Imagen_Login.png").resize((210, 220))
            self.render = ImageTk.PhotoImage(self.imagen)
            Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        except Exception:
            Label(self.parte_superior, text="[Imagen login no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray").pack(expand=True, fill="both", side="top")

        # Parte inferior: entradas y botones
        self.parte_inferior = Frame(self.derecha, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)

        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        # Entradas
        Label(self.parte_inferior, text="Usuario", font=("Arial", 18), bg=fondo).grid(row=0, column=0, padx=10, sticky="e")
        self.entry_usuario = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18))
        self.entry_usuario.grid(row=0, column=1, columnspan=3, padx=5, sticky="w")

        Label(self.parte_inferior, text="Contraseña:", font=("Arial", 18), bg=fondo).grid(row=1, column=0, pady=10, sticky="e")
        self.entry_contrasena = Entry(self.parte_inferior, bd=0, width=14, font=("Arial", 18), show="*")
        self.entry_contrasena.grid(row=1, column=1, columnspan=3, padx=5, sticky="w")

        # Botones
        Button(self.parte_inferior, text="Ingresar", width=16, font=("Arial", 12),bg= color_botones, command=self.verificacion_datos, relief="raised", bd=4).grid(row=2, column=0, columnspan=2, padx=35, pady=5)

        Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),bg= color_botones, command=lambda: self.regresar(ventana), relief="raised", bd=4).grid(row=8, column=0, columnspan=4, padx=35, pady=5)

    #Método para redimensionar la imagen izquierda
    def redimensionar_imagen(self, event):
        nueva_img = self.original.resize((event.width, event.height))
        self.img = ImageTk.PhotoImage(nueva_img)
        self.label_imagen.config(image=self.img)

    #Regresar al Login inicial
    def regresar(self, ventana):
        self.ventana.destroy()
        ventana.deiconify()

    #Verificación de usuario y contraseña
    def verificacion_datos(self):
        usuario = self.entry_usuario.get()
        contrasena = self.entry_contrasena.get()

        if not usuario or not contrasena:
            messagebox.showwarning("Cuidado!", "Se deben completar ambos campos")
            return

        if not os.path.exists(archivo_usuarios):
            messagebox.showerror("Error", "No hay usuarios registrados")
            return

        # Acceso especial administrador
        if self.tipo == "Administrador" and usuario == "admin" and contrasena == "123":
            messagebox.showinfo("Éxito", "Bienvenido admin como Administrador")
            self.ventana.destroy()
            Restaurante(self.ventana.master, self.tipo, usuario)
            return

        wb_temp = load_workbook(archivo_usuarios)
        ws_temp = wb_temp.active

        encontrado = False
        for fila in ws_temp.iter_rows(values_only=True):
            if fila[0] == usuario and fila[1] == contrasena and fila[2] == self.tipo:
                encontrado = True
                break

        if encontrado:
            messagebox.showinfo("Éxito", f"Bienvenido {usuario} como {self.tipo}")
            self.ventana.destroy()
            Restaurante(self.ventana.master, self.tipo, usuario)
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrecto")