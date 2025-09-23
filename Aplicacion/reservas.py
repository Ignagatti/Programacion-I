# reservas.py
from tkinter import Toplevel, Frame, Button, Label, Entry, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook
import os

class Reservar_mesa:
    def __init__(self, ventana, tipo, restaurante, mesa_id):
        #constructores
        self.tipo = tipo
        self.restaurante = restaurante
        self.mesa_id = mesa_id

        self.ventana = Toplevel(ventana)
        self.ventana.geometry("400x700")
        self.ventana.title("Reservar Mesa")

        fondo = "#588E6B"
        color_botones= "#A7CBBF"

        #Creamos los frames superior e inferior
        self.parte_superior = Frame(self.ventana, bg=fondo)
        self.parte_superior.pack(fill="both", expand=True)

        self.parte_inferior = Frame(self.ventana, bg=fondo)
        self.parte_inferior.pack(fill="both", expand=True)
        self.parte_inferior.columnconfigure(0, weight=1)
        self.parte_inferior.columnconfigure(1, weight=1)

        # Recuperar capacidad
        capacidad = self.restaurante.estado_mesas[self.mesa_id].get("capacidad", "4")
        titulo = (
            f"Admin - Mesa {mesa_id} (Capacidad: {capacidad})"
            if self.tipo == "Administrador"
            else f"Reservar Mesa {mesa_id} (Capacidad: {capacidad})"
        )
        Label(self.parte_superior, text=titulo, font=("Calisto MT", 20, "bold"), bg=fondo).pack(pady=20)

       

        mesa_info = self.restaurante.estado_mesas[self.mesa_id]

        # Restricción si está ocupada
        if mesa_info["estado"] == "Ocupada" and self.tipo != "Administrador" and mesa_info["nombre"] != self.restaurante.usuario:
            Label(
                self.parte_inferior,
                text=f"Mesa ocupada por {mesa_info['nombre']}",
                font=("Arial", 16), bg=fondo
            ).grid(row=0, column=0, columnspan=2, pady=10)
            Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),bg=color_botones,
                   command=self.regresar).grid(row=2, column=0, columnspan=2, pady=5)
            return

        # Botones según el tipo de usuario
        if self.tipo == "Administrador":
            Button(self.parte_inferior, text="Desocupar", width=16, font=("Arial", 12),bg=color_botones,
                   command=self.desocupar).grid(row=2, column=0, columnspan=2, pady=5)
            
            Button(self.parte_inferior, text="Eliminar", width=16, font=("Arial", 12),bg=color_botones,
                   command=self.eliminar).grid(row=4, column=0, columnspan=2, pady=5)
            
            Button(self.parte_inferior, text="Modificar Capacidad", width=16, font=("Arial", 12),bg=color_botones,
                   command=self.modificar_capacidad).grid(row=5, column=0, columnspan=2, pady=5)
            
            Button(self.parte_inferior, text="Modificar Número", width=16, font=("Arial", 12),bg=color_botones,
                    command=self.modificar_numero_mesa).grid(row=6, column=0, columnspan=2, pady=5)
            
            Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),bg=color_botones,
                   command=self.regresar).grid(row=7, column=0, columnspan=2, pady=5)
        else:
            if mesa_info["estado"] == "Libre":
                Button(self.parte_inferior, text="Reservar", width=16, font=("Arial", 12),bg=color_botones, command=self.guardar).grid(row=2, column=0, columnspan=2, pady=5)

            Button(self.parte_inferior, text="Atrás", width=16, font=("Arial", 12),bg=color_botones,command=self.regresar).grid(row=4, column=0, columnspan=2, pady=5)

        # Mostrar botón cancelar si corresponde
        self.mostrar_cancelar()
    
     # Imagen
        try:
            if mesa_info["estado"] == "Ocupada":
                self.imagen= Image.open("reservado.png").resize((210,220))
                self.render = ImageTk.PhotoImage(self.imagen)
                Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
            else:
                self.imagen = Image.open("reserva.png").resize((210, 220))
                self.render = ImageTk.PhotoImage(self.imagen)
                Label(self.parte_superior, image=self.render, bg=fondo).pack(expand=True, fill="both", side="top")
        except Exception:
            Label(self.parte_superior, text="[Imagen reserva no encontrada]", bg=fondo, font=("Arial", 14, "italic"), fg="gray").pack(expand=True, fill="both", side="top")

    #MÉTODOS 

    def regresar(self):
        self.ventana.destroy()

    def guardar(self):
        #Guardar la reserva hecha por un usuario
        nombre = self.restaurante.usuario  # Usar el usuario logueado como nombre de reserva
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["estado"] == "Libre":
            capacidad = mesa_info.get("capacidad", "4")
            self.restaurante.estado_mesas[self.mesa_id] = {"estado": "Ocupada", "nombre": nombre, "capacidad": capacidad}
            self.restaurante.botones_mesas[self.mesa_id].configure(
                bg="red", text=f"Mesa {self.mesa_id}\n({nombre})\nCapacidad: {capacidad}"
            )
            self.actualizar_excel("Ocupada", nombre, capacidad)
            self.ventana.destroy()
        else:
            messagebox.showerror("Error", "Esta mesa ya está ocupada")

    def modificar_capacidad(self):
        ventana_cap = Toplevel(self.ventana)
        ventana_cap.title("Modificar Capacidad")
        ventana_cap.geometry("300x150")

        Label(ventana_cap, text="Nueva capacidad:", font=("Arial", 14)).pack(pady=10)
        entry_cap = Entry(ventana_cap, font=("Arial", 14))
        entry_cap.pack(pady=5)

        def guardar():
            nueva_cap = entry_cap.get()
            if not nueva_cap.isdigit() or int(nueva_cap) <= 0:
                messagebox.showerror("Error", "Ingrese un número válido")
                return

            mesa_info = self.restaurante.estado_mesas[self.mesa_id]
            mesa_info["capacidad"] = nueva_cap

            texto = f"Mesa {self.mesa_id}\nCapacidad: {nueva_cap}" if mesa_info["estado"] == "Libre" else f"Mesa {self.mesa_id}\n({mesa_info['nombre']})\nCapacidad: {nueva_cap}"
            self.restaurante.botones_mesas[self.mesa_id].configure(text=texto)

            self.actualizar_excel(mesa_info["estado"], mesa_info["nombre"], nueva_cap)
            ventana_cap.destroy()

        Button(ventana_cap, text="Guardar", command=guardar).pack(pady=10)

    def modificar_numero_mesa(self):
        ventana_num = Toplevel(self.ventana)
        ventana_num.title("Modificar Número de Mesa")
        ventana_num.geometry("300x150")

        Label(ventana_num, text="Nuevo número de mesa:", font=("Arial", 14)).pack(pady=10)
        entry_num = Entry(ventana_num, font=("Arial", 14))
        entry_num.pack(pady=5)

        def guardar():
            nuevo_id = entry_num.get()
            if not nuevo_id.isdigit() or int(nuevo_id) <= 0:
                messagebox.showerror("Error", "Ingrese un número válido")
                return
            nuevo_id = int(nuevo_id)

            # Verificar que no exista ya ese número
            if nuevo_id in self.restaurante.estado_mesas:
                messagebox.showerror("Error", "Ya existe una mesa con ese número")
                return

            mesa_info = self.restaurante.estado_mesas.pop(self.mesa_id)
            btn = self.restaurante.botones_mesas.pop(self.mesa_id)

            # Actualizar diccionarios con el nuevo ID
            self.restaurante.estado_mesas[nuevo_id] = mesa_info
            self.restaurante.botones_mesas[nuevo_id] = btn

            # Actualizar texto del botón
            capacidad = mesa_info.get("capacidad", "4")
            texto = f"Mesa {nuevo_id}\nCapacidad: {capacidad}" if mesa_info["estado"] == "Libre" else f"Mesa {nuevo_id}\n({mesa_info['nombre']})\nCapacidad: {capacidad}"
            btn.configure(text=texto, command=lambda n=nuevo_id: Reservar_mesa(self.restaurante.ventana, self.tipo, self.restaurante, n))

            # Actualizar Excel
            archivo = "Datos_Reservas.xlsx"
            reservas = load_workbook(archivo)
            hoja = reservas.active
            for fila in hoja.iter_rows(min_row=2):
                if fila[0].value == self.mesa_id:
                    fila[0].value = nuevo_id
                    break
            reservas.save(archivo)

            self.mesa_id = nuevo_id
            ventana_num.destroy()

        Button(ventana_num, text="Guardar", command=guardar).pack(pady=10)

    def eliminar(self):
        if messagebox.askyesno("Eliminar Mesa", "¿Está seguro que desea eliminar esta mesa?"):
            self.eliminar_de_excel()
            btn = self.restaurante.botones_mesas.get(self.mesa_id)
            if btn:
                btn.destroy()
                del self.restaurante.botones_mesas[self.mesa_id]
            self.restaurante.estado_mesas.pop(self.mesa_id, None)
            self.ventana.destroy()

    def desocupar(self):
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["nombre"] == self.restaurante.usuario or self.tipo == "Administrador":
            capacidad = mesa_info.get("capacidad", "4")
            self.restaurante.estado_mesas[self.mesa_id] = {"estado": "Libre", "nombre": "", "capacidad": capacidad}
            self.restaurante.botones_mesas[self.mesa_id].configure(
                bg="green", text=f"Mesa {self.mesa_id}\nCapacidad: {capacidad}"
            )
            self.actualizar_excel("Libre", "", capacidad)
            self.ventana.destroy()
        else:
            messagebox.showerror("Error", "Solo el administrador o el usuario que reservó puede desocupar la mesa.")

    def actualizar_excel(self, estado, nombre="", capacidad="4"):
        archivo = "Datos_Reservas.xlsx"
        reservas = load_workbook(archivo)
        hoja = reservas.active

        for fila in hoja.iter_rows(min_row=2):
            if fila[0].value == self.mesa_id:
                fila[1].value = nombre
                fila[2].value = estado
                fila[3].value = capacidad
                reservas.save(archivo)
                return

        hoja.append([self.mesa_id, nombre, estado, capacidad])
        reservas.save(archivo)

    
    def cancelar_reserva(self):
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["estado"] == "Ocupada" and mesa_info["nombre"] == self.restaurante.usuario:
            capacidad = mesa_info.get("capacidad", "4")
            self.restaurante.estado_mesas[self.mesa_id] = {"estado": "Libre", "nombre": "", "capacidad": capacidad}
            self.restaurante.botones_mesas[self.mesa_id].configure(
                bg="green", text=f"Mesa {self.mesa_id}\nCapacidad: {capacidad}"
            )
            self.actualizar_excel("Libre", "", capacidad)
            self.ventana.destroy()
        else:
            messagebox.showerror("Error", "Solo el usuario que reservó puede cancelar la reserva.")

    def mostrar_cancelar(self):
        color_botones= "#A7CBBF"
        mesa_info = self.restaurante.estado_mesas[self.mesa_id]
        if mesa_info["estado"] == "Ocupada" and mesa_info["nombre"] == self.restaurante.usuario and self.tipo != "Administrador":
            Button(self.parte_inferior, text="Cancelar Reserva", width=16, font=("Arial", 12),bg=color_botones,
                   command=self.cancelar_reserva).grid(row=2, column=0, columnspan=2, pady=5)

    def eliminar_de_excel(self):
        archivo = "Datos_Reservas.xlsx"
        reservas = load_workbook(archivo)
        hoja = reservas.active

        for fila in hoja.iter_rows(min_row=2):
            if fila[0].value == self.mesa_id:
                hoja.delete_rows(fila[0].row)
                break

        reservas.save(archivo)


